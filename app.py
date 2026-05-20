import streamlit as st
import pandas as pd
from api.alphavantage import get_news_sentiment, get_quote
from api.mediastack import get_india_news
from api.portfolio import PORTFOLIO, THEMES
from api.nifty50 import NIFTY_50
from api.glm5 import analyze_stock, analyze_theme, stream_chat
from api.angelone import fetch_all, is_configured

st.set_page_config(page_title="Stock Market Analyser", page_icon="📈", layout="wide")
st.title("📈 Stock Market Analyser — India Focus")

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "🗂 My Portfolio", "🌐 Theme Analysis",
    "🔍 Nifty 50 News", "🤖 AI Analysis", "💹 Angel One", "🤖 Trading Bot",
])


# ── Tab 1: My Portfolio ────────────────────────────────────────────────────────
with tab1:
    st.subheader("My Portfolio — News & Market Impact")
    st.caption("News from Mediastack · Sentiment & price from AlphaVantage")

    for name, info in PORTFOLIO.items():
        ticker = info["ticker"]
        with st.expander(f"**{name}** ({ticker})  |  Sector: {info['sector']}", expanded=True):
            col_news, col_price = st.columns([2, 1])

            with col_price:
                st.markdown("**Live Quote**")
                with st.spinner("Fetching price..."):
                    quote = get_quote(ticker)
                if "error" in quote:
                    st.warning(quote["error"])
                else:
                    change_pct = quote["change_percent"].replace("%", "").strip()
                    try:
                        val = float(change_pct)
                        color = "🟢" if val >= 0 else "🔴"
                        st.metric(
                            label=f"{ticker}",
                            value=f"₹ {float(quote['price']):.2f}",
                            delta=f"{quote['change_percent']}",
                        )
                    except ValueError:
                        st.write(f"Price: {quote['price']}")

                st.markdown("**AlphaVantage Sentiment**")
                with st.spinner("Fetching sentiment..."):
                    av = get_news_sentiment(ticker)
                if "error" in av:
                    st.warning(av["error"])
                elif av.get("data"):
                    scores = []
                    for a in av["data"][:5]:
                        score = a.get("Score", 0)
                        try:
                            scores.append(float(score))
                        except (ValueError, TypeError):
                            pass
                    if scores:
                        avg = sum(scores) / len(scores)
                        label = "Bullish 🐂" if avg > 0.15 else "Bearish 🐻" if avg < -0.15 else "Neutral ⚖️"
                        st.metric("Avg Sentiment Score", f"{avg:.3f}", delta=label)
                else:
                    st.info("No AV sentiment data.")

            with col_news:
                st.markdown("**Latest News (Mediastack)**")
                with st.spinner("Fetching news..."):
                    news = get_india_news(info["keywords"], limit=6)
                if "error" in news:
                    st.warning(news["error"])
                elif news.get("data"):
                    for article in news["data"]:
                        st.markdown(
                            f"- [{article['title']}]({article['url']})  \n"
                            f"  <small>📰 {article['source']} &nbsp;·&nbsp; {article['published_at'][:10] if article['published_at'] else ''}</small>",
                            unsafe_allow_html=True,
                        )
                else:
                    st.info("No news found for this stock.")


# ── Tab 2: Theme Analysis ──────────────────────────────────────────────────────
with tab2:
    st.subheader("Macro Theme → Portfolio Impact")
    st.caption("Select a global/macro event to see how it affects your holdings")

    theme_name = st.selectbox("Select a Theme / Macro Event", list(THEMES.keys()))
    theme = THEMES[theme_name]

    st.info(f"**Impact logic:** {theme['impact_note']}")

    col_left, col_right = st.columns([1, 1])

    with col_left:
        st.markdown("**Related News Headlines** (Mediastack)")
        with st.spinner("Fetching theme news..."):
            news = get_india_news(theme["keywords"], limit=8)
        if "error" in news:
            st.warning(news["error"])
        elif news.get("data"):
            for article in news["data"]:
                sentiment_hint = ""
                title_lower = article["title"].lower()
                if any(w in title_lower for w in ["rise", "surge", "gain", "up", "boost", "high"]):
                    sentiment_hint = "🟢 "
                elif any(w in title_lower for w in ["fall", "drop", "down", "loss", "weak", "cut"]):
                    sentiment_hint = "🔴 "
                st.markdown(
                    f"{sentiment_hint}[{article['title']}]({article['url']})  \n"
                    f"<small>{article['source']} · {article['published_at'][:10] if article['published_at'] else ''}</small>",
                    unsafe_allow_html=True,
                )
        else:
            st.info("No news found for this theme.")

    with col_right:
        st.markdown("**Affected Portfolio Stocks**")
        affected = {
            name: info
            for name, info in PORTFOLIO.items()
            if info["sector"] in theme["affected_sectors"]
        }
        if not affected:
            st.info("None of your portfolio stocks are in the affected sectors for this theme.")
        else:
            rows = []
            for name, info in affected.items():
                quote = get_quote(info["ticker"])
                if "error" not in quote:
                    rows.append({
                        "Stock": name,
                        "Ticker": info["ticker"],
                        "Price (₹)": quote["price"],
                        "Change": quote["change_percent"],
                    })
                else:
                    rows.append({
                        "Stock": name,
                        "Ticker": info["ticker"],
                        "Price (₹)": "—",
                        "Change": "—",
                    })

            df = pd.DataFrame(rows)

            def color_change(val):
                try:
                    v = float(str(val).replace("%", "").strip())
                    return "color: green" if v >= 0 else "color: red"
                except ValueError:
                    return ""

            st.dataframe(
                df.style.map(color_change, subset=["Change"]),
                use_container_width=True,
                hide_index=True,
            )

            st.markdown("**News for Affected Stocks**")
            for name, info in affected.items():
                with st.expander(name):
                    n = get_india_news(info["keywords"], limit=4)
                    if n.get("data"):
                        for a in n["data"]:
                            st.markdown(f"- [{a['title']}]({a['url']})", unsafe_allow_html=True)
                    else:
                        st.info("No recent news.")


# ── Tab 3: Nifty 50 News ───────────────────────────────────────────────────────
with tab3:
    st.subheader("Nifty 50 — News & Sentiment")

    col_s, col_m = st.columns([1, 2])
    with col_s:
        mode = st.radio("Select by", ["Nifty 50 Dropdown", "Custom Ticker"], horizontal=True)
        if mode == "Nifty 50 Dropdown":
            company = st.selectbox("Stock", list(NIFTY_50.keys()))
            ticker = NIFTY_50[company]
            st.caption(f"Ticker: `{ticker}`")
        else:
            ticker = st.text_input("Enter BSE ticker (e.g. RELIANCE.BSE)", value="RELIANCE.BSE")

        fetch = st.button("Analyse", type="primary")

    if fetch:
        with col_m:
            st.markdown(f"### {ticker}")
            q = get_quote(ticker)
            if "error" not in q:
                st.metric("Price", f"₹ {float(q['price']):.2f}", delta=q["change_percent"])

            st.markdown("**News (Mediastack)**")
            kw = [ticker.split(".")[0]]
            mn = get_india_news(kw, limit=8)
            if mn.get("data"):
                for a in mn["data"]:
                    st.markdown(f"- [{a['title']}]({a['url']})", unsafe_allow_html=True)
            else:
                st.info("No Mediastack news found.")

            st.markdown("**Sentiment (AlphaVantage)**")
            av = get_news_sentiment(ticker)
            if av.get("data"):
                df = pd.DataFrame(av["data"])
                st.dataframe(
                    df[["Headline", "Sentiment", "Score", "Source", "Published"]],
                    use_container_width=True,
                    hide_index=True,
                )
            else:
                st.info("No AlphaVantage sentiment data.")


# ── Tab 4: AI Analysis ─────────────────────────────────────────────────────────
with tab4:
    st.subheader("🤖 GLM-5 AI Analysis")
    st.caption("Powered by Llama 3.3 70B · via OpenRouter · Free model · No local setup needed")

    ai_tab1, ai_tab2, ai_tab3 = st.tabs(["📊 Portfolio Stock Analysis", "🌐 Theme Analysis", "💬 Ask AI"])

    # ── AI Sub-tab 1: Portfolio stock analysis ──────────────────────────────────
    with ai_tab1:
        st.markdown("Select a stock from your portfolio to get a GLM-5 generated analysis based on live news and sentiment.")

        stock_choice = st.selectbox("Choose a portfolio stock", list(PORTFOLIO.keys()), key="ai_stock_select")

        if st.button("Generate AI Analysis", type="primary", key="ai_stock_btn"):
            info = PORTFOLIO[stock_choice]
            ticker = info["ticker"]

            with st.spinner(f"Fetching data for {stock_choice}..."):
                news_data = get_india_news(info["keywords"], limit=6)
                av_data = get_news_sentiment(ticker)

            news_list = news_data.get("data", [])
            sentiment_list = av_data.get("data", [])

            with st.spinner("GLM-5 is analysing..."):
                result = analyze_stock(stock_choice, ticker, news_list, sentiment_list)

            st.markdown(f"### Analysis: {stock_choice} ({ticker})")
            st.markdown(result)

    # ── AI Sub-tab 2: Theme AI analysis ────────────────────────────────────────
    with ai_tab2:
        st.markdown("Select a macro theme to get GLM-5's analysis of how it affects your portfolio.")

        theme_choice = st.selectbox("Choose a macro theme", list(THEMES.keys()), key="ai_theme_select")

        if st.button("Generate Theme Analysis", type="primary", key="ai_theme_btn"):
            theme = THEMES[theme_choice]

            with st.spinner(f"Fetching news for theme: {theme_choice}..."):
                theme_news = get_india_news(theme["keywords"], limit=8)

            affected = [
                name for name, info in PORTFOLIO.items()
                if info["sector"] in theme["affected_sectors"]
            ]

            with st.spinner("GLM-5 is analysing the macro theme..."):
                result = analyze_theme(
                    theme_choice,
                    theme["impact_note"],
                    theme_news.get("data", []),
                    affected,
                )

            st.markdown(f"### Macro Theme: {theme_choice}")
            if affected:
                st.caption(f"Affected portfolio stocks: {', '.join(affected)}")
            st.markdown(result)

    # ── AI Sub-tab 3: Free-form chat ────────────────────────────────────────────
    with ai_tab3:
        st.markdown("Ask GLM-5 anything about your portfolio, the Indian market, or investing strategy.")

        # Build a brief portfolio context string to inject into every message
        portfolio_context = "\n".join(
            f"- {name} ({info['ticker']}, sector: {info['sector']})"
            for name, info in PORTFOLIO.items()
        )

        if "chat_history" not in st.session_state:
            st.session_state.chat_history = []

        # Display chat history
        for msg in st.session_state.chat_history:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        user_input = st.chat_input("Ask about your portfolio or the Indian market...")

        if user_input:
            st.session_state.chat_history.append({"role": "user", "content": user_input})
            with st.chat_message("user"):
                st.markdown(user_input)

            system_msg = {
                "role": "system",
                "content": (
                    "You are a financial assistant specializing in Indian stock markets. "
                    f"The user's portfolio:\n{portfolio_context}\n"
                    "Answer concisely and specifically. Focus on Indian market dynamics."
                ),
            }

            messages_to_send = [system_msg] + st.session_state.chat_history

            with st.chat_message("assistant"):
                response_placeholder = st.empty()
                full_response = ""
                for chunk in stream_chat(messages_to_send):
                    full_response += chunk
                    response_placeholder.markdown(full_response + "▌")
                response_placeholder.markdown(full_response)

            st.session_state.chat_history.append({"role": "assistant", "content": full_response})

        if st.session_state.chat_history:
            if st.button("Clear Chat", key="clear_chat"):
                st.session_state.chat_history = []
                st.rerun()


# ── Tab 5: Angel One Live Data ─────────────────────────────────────────────────
with tab5:
    st.subheader("💹 Angel One — Live Trading Data")

    # ── Config check ───────────────────────────────────────────────────────────
    if not is_configured():
        st.error("Angel One credentials are not configured.")
        st.markdown("""
**Steps to set up:**

1. Create an app at [smartapi.angelone.in](https://smartapi.angelone.in) and copy your **API Key**
2. Open `.env` in this project and fill in:
```
ANGELONE_API_KEY=your_api_key
ANGELONE_CLIENT_CODE=your_client_code   # e.g. R123456
ANGELONE_PASSWORD=your_pin              # 4-digit trading PIN
ANGELONE_TOTP_SECRET=your_totp_secret   # base32 secret from Angel One TOTP setup
```
3. To get the **TOTP secret**: Angel One App → My Profile → Account Security → Enable TOTP
   → tap "Can't scan QR?" → copy the base32 secret shown
4. Restart this app after editing `.env`
        """)
        st.stop()

    # ── Fetch button ───────────────────────────────────────────────────────────
    col_btn, col_status = st.columns([1, 3])
    with col_btn:
        fetch_ao = st.button("🔄 Fetch from Angel One", type="primary", key="ao_fetch")
    with col_status:
        if "ao_fetched" in st.session_state:
            st.caption(f"Last fetched at {st.session_state.get('ao_fetch_time', '—')}")

    if fetch_ao:
        import datetime
        with st.spinner("Fetching data from Angel One…"):
            result = fetch_all()
            st.session_state["ao_profile"]   = result["profile"]
            st.session_state["ao_funds"]     = result["funds"]
            st.session_state["ao_trades"]    = result["trades"]
            st.session_state["ao_orders"]    = result["orders"]
            st.session_state["ao_holdings"]  = result["holdings"]
            st.session_state["ao_positions"] = result["positions"]
            st.session_state["ao_fetched"]   = True
            st.session_state["ao_fetch_time"] = datetime.datetime.now().strftime("%H:%M:%S")

        first_error = result["profile"]["error"]
        if first_error:
            st.error(f"Error: {first_error}")
            _hint = ""
            if "base32" in first_error.lower() or "totp" in first_error.lower():
                _hint = "**TOTP secret issue:** Check `ANGELONE_TOTP_SECRET` in `.env` — must be A-Z and 2-7 only."
            elif "rate" in first_error.lower() or "access" in first_error.lower():
                _hint = "**Rate limit:** Angel One blocked too many logins. Wait 5–10 minutes and try again. The token cache has been cleared."
            elif "invalid token" in first_error.lower() or "ag8001" in first_error.lower():
                _hint = "**Session expired:** Angel One token is no longer valid. The cache has been cleared — click **Fetch from Angel One** again to log in fresh."
            elif "login failed" in first_error.lower():
                _hint = "**Login failed:** Double-check `ANGELONE_CLIENT_CODE`, `ANGELONE_PASSWORD`, and `ANGELONE_API_KEY`."
            if _hint:
                st.info(_hint)
        else:
            from_cache = result["login"].get("from_cache", False)
            st.success(f"Data fetched successfully. {'(used cached token — no re-login)' if from_cache else '(logged in fresh)'}")

    if "ao_fetched" not in st.session_state:
        st.info("Click **Fetch from Angel One** to load your live trading data.")
        st.stop()

    # ── Helper: error banner ───────────────────────────────────────────────────
    def _ao_error(result: dict, label: str):
        if not result["status"]:
            st.error(f"{label} error: {result['error']}")
            return True
        return False

    # ── Helper: color P&L cells ────────────────────────────────────────────────
    def _color_pnl(val):
        try:
            return "color: #22c55e" if float(val) >= 0 else "color: #ef4444"
        except (TypeError, ValueError):
            return ""

    # ── Sub-tabs ───────────────────────────────────────────────────────────────
    ao1, ao2, ao3, ao4, ao5 = st.tabs([
        "👤 Account & Funds",
        "📋 Trade Book",
        "📒 Order Book",
        "🏦 Holdings",
        "📊 Positions",
    ])

    # ── AO Sub-tab 1: Account & Funds ──────────────────────────────────────────
    with ao1:
        profile_res = st.session_state["ao_profile"]
        funds_res   = st.session_state["ao_funds"]

        col_p, col_f = st.columns(2)

        with col_p:
            st.markdown("#### Profile")
            if not _ao_error(profile_res, "Profile"):
                p = profile_res["data"]
                st.markdown(f"**Name:** {p.get('name', '—')}")
                st.markdown(f"**Email:** {p.get('email', '—')}")
                st.markdown(f"**Mobile:** {p.get('mobileNo', '—')}")
                st.markdown(f"**Client Code:** {p.get('clientcode', '—')}")
                exchanges = p.get('exchanges', [])
                products  = p.get('products', [])
                if exchanges:
                    st.markdown(f"**Exchanges enabled:** {', '.join(exchanges)}")
                if products:
                    st.markdown(f"**Products enabled:** {', '.join(products)}")

        with col_f:
            st.markdown("#### Funds & Margins")
            if funds_res["status"] is False:
                st.error(f"Funds error: {funds_res['error']}")
            else:
                f = funds_res["data"] or {}
                def _fval(key):
                    v = f.get(key, "")
                    return v if v not in ("", None) else "—"

                fund_rows = [
                    ("Net Available (₹)",          _fval("net")),
                    ("Available Cash (₹)",          _fval("availablecash")),
                    ("Available Cash Margin (₹)",   _fval("availablecashmargain")),
                    ("Collateral (₹)",              _fval("collateral")),
                    ("Utilised Debits (₹)",         _fval("utiliseddebits")),
                    ("M2M Realised (₹)",            _fval("m2mrealized")),
                    ("M2M Unrealised (₹)",          _fval("m2munrealized")),
                ]

                if not f:
                    st.warning("Angel One returned no funds data. This sometimes happens with a cached token — click **Fetch from Angel One** to force a fresh login.")
                elif all(v == "—" for _, v in fund_rows):
                    # Fields exist but use different key names — show raw
                    st.caption("Showing raw fields from API (field names differ from expected):")
                    fund_rows = [(k, str(v)) for k, v in f.items() if v not in ("", None)]

                if fund_rows:
                    funds_df = pd.DataFrame(fund_rows, columns=["Item", "Value"])
                    st.dataframe(funds_df, use_container_width=True, hide_index=True)

                # Always show raw debug, expanded when data is missing
                with st.expander("Raw API response (debug)", expanded=not bool(f)):
                    st.json(f)

    # ── AO Sub-tab 2: Trade Book ───────────────────────────────────────────────
    with ao2:
        st.markdown("#### Today's Executed Trades")
        st.caption("Only filled orders appear here. Refreshes with each Fetch.")
        trades_res = st.session_state["ao_trades"]
        if not _ao_error(trades_res, "Trade Book"):
            data = trades_res["data"]
            if not data:
                st.info("No executed trades today.")
            else:
                raw_df = pd.DataFrame(data)

                # Angel One uses "quantity" or "qty" depending on API version
                def _pick(df, *candidates):
                    for c in candidates:
                        if c in df.columns:
                            return c
                    return None

                qty_col  = _pick(raw_df, "quantity", "qty", "Quantity")
                px_col   = _pick(raw_df, "price", "Price", "tradeprice")
                val_col  = _pick(raw_df, "tradevalue", "tradeValue", "value")
                side_col = _pick(raw_df, "transactiontype", "side", "Side")
                sym_col  = _pick(raw_df, "tradingsymbol", "symbol")

                TRADE_COLS = {}
                if sym_col:  TRADE_COLS[sym_col]  = "Symbol"
                col_map = {
                    _pick(raw_df, "exchange"):      "Exchange",
                    side_col:                       "Side",
                    _pick(raw_df, "producttype"):   "Product",
                    qty_col:                        "Qty",
                    px_col:                         "Price ₹",
                    val_col:                        "Value ₹",
                    _pick(raw_df, "tradetime", "tradeTime"): "Trade Time",
                    _pick(raw_df, "orderid", "orderId"):     "Order ID",
                }
                for k, v in col_map.items():
                    if k:
                        TRADE_COLS[k] = v

                present_cols = [k for k in TRADE_COLS if k in raw_df.columns]
                df = raw_df[present_cols].rename(columns=TRADE_COLS)

                for col in ["Qty", "Price ₹", "Value ₹"]:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce")

                def _color_side(val):
                    if str(val).upper() == "BUY":  return "color: #22c55e"
                    if str(val).upper() == "SELL": return "color: #ef4444"
                    return ""

                styled = df.style
                if "Side" in df.columns:
                    styled = styled.map(_color_side, subset=["Side"])
                num_fmt = {c: "{:.2f}" for c in ["Price ₹", "Value ₹"] if c in df.columns}
                if num_fmt:
                    styled = styled.format(num_fmt)

                st.dataframe(styled, use_container_width=True, hide_index=True)
                st.caption(f"{len(df)} trade(s) executed today")

                with st.expander("Raw field names (debug — check if Qty/Price missing)"):
                    st.write(list(raw_df.columns))

                # ── P&L breakdown for SELL trades ─────────────────────────────
                sell_trades = df[df["Side"].str.upper() == "SELL"] if "Side" in df.columns else pd.DataFrame()
                if not sell_trades.empty:
                    st.markdown("---")
                    st.markdown("#### 💰 Today's Realised P&L (SELL trades)")
                    st.caption("Avg buy price sourced from your Angel One holdings (preferred) or stock_config.xlsx.")

                    # ── Build avg buy price map: prefer live holdings from session state ──
                    avg_buy_map = {}

                    # Source 1: Holdings already fetched in this session
                    holdings_data = (st.session_state.get("ao_holdings") or {}).get("data") or []
                    for h in holdings_data:
                        raw_sym = str(h.get("tradingsymbol") or "").upper().replace("-EQ", "").strip()
                        try:
                            avg = float(h.get("averageprice") or 0)
                            if avg > 0:
                                avg_buy_map[raw_sym] = avg
                        except (ValueError, TypeError):
                            pass

                    # Source 2: Excel fallback for any still missing
                    try:
                        import openpyxl as _opxl
                        _wb = _opxl.load_workbook("stock_config.xlsx", read_only=True, data_only=True)
                        if "My Holdings" in _wb.sheetnames:
                            _ws = _wb["My Holdings"]
                            _hdr = {str(_ws.cell(1, c).value or "").strip().lower(): c
                                    for c in range(1, _ws.max_column + 1)}
                            _sym_c = next((v for k, v in _hdr.items() if "symbol" in k), None)
                            _avg_c = next((v for k, v in _hdr.items() if "avg" in k or "buy price" in k), None)
                            if _sym_c and _avg_c:
                                for _r in range(2, _ws.max_row + 1):
                                    _s = str(_ws.cell(_r, _sym_c).value or "").strip().upper()
                                    _a = _ws.cell(_r, _avg_c).value
                                    try:
                                        if _s and _a and _s not in avg_buy_map:
                                            avg_buy_map[_s] = float(_a)
                                    except (ValueError, TypeError):
                                        pass
                        _wb.close()
                    except Exception:
                        pass

                    # Build Order Book lookup: symbol → {qty, price} for completed SELL orders
                    # Trade Book often omits qty/price; Order Book always has them.
                    ob_lookup: dict[str, dict] = {}
                    ob_data = (st.session_state.get("ao_orders") or {}).get("data") or []
                    for ob_row in ob_data:
                        ob_sym  = str(ob_row.get("tradingsymbol") or "").upper().replace("-EQ", "").strip()
                        ob_side = str(ob_row.get("transactiontype") or "").upper()
                        ob_stat = str(ob_row.get("orderstatus") or "").upper()
                        if ob_side == "SELL" and "COMPLETE" in ob_stat:
                            try:
                                ob_lookup[ob_sym] = {
                                    "qty": float(ob_row.get("quantity") or ob_row.get("qty") or 0),
                                    "px":  float(ob_row.get("price") or ob_row.get("averageprice") or 0),
                                }
                            except (ValueError, TypeError):
                                pass

                    def _num(v):
                        """Return float if v is a usable number, else None."""
                        try:
                            f = float(v)
                            return f if f > 0 else None
                        except (TypeError, ValueError):
                            return None

                    pnl_rows = []
                    for _, tr in sell_trades.iterrows():
                        raw_sym  = str(tr.get("Symbol", "")).upper().replace("-EQ", "").strip()
                        sell_val = _num(tr.get("Value ₹"))
                        qty      = _num(tr.get("Qty"))
                        sell_px  = _num(tr.get("Price ₹"))

                        # Fall back to Order Book when Trade Book is missing qty/price
                        if qty is None or sell_px is None:
                            ob = ob_lookup.get(raw_sym, {})
                            if qty is None:     qty     = _num(ob.get("qty"))
                            if sell_px is None: sell_px = _num(ob.get("px"))

                        # Derive sell_val from qty × price if still missing
                        if sell_val is None and qty and sell_px:
                            sell_val = round(qty * sell_px, 2)

                        avg_buy = avg_buy_map.get(raw_sym)
                        if avg_buy and qty and sell_px:
                            sell_val = sell_val or round(qty * sell_px, 2)
                            buy_val  = round(qty * avg_buy, 2)
                            pnl      = round(sell_val - buy_val, 2)
                            pnl_pct  = round(pnl / buy_val * 100, 2) if buy_val else None
                        else:
                            buy_val = pnl = pnl_pct = None

                        pnl_rows.append({
                            "Stock (Symbol)":  raw_sym,
                            "Qty Sold":        qty,
                            "Sell Price ₹":    sell_px,
                            "Sell Value ₹":    sell_val,
                            "Avg Buy Price ₹": avg_buy,
                            "Buy Cost ₹":      buy_val,
                            "Profit / Loss ₹": pnl,
                            "Return %":        pnl_pct,
                        })

                    pnl_df = pd.DataFrame(pnl_rows)

                    def _pnl_color(val):
                        try:
                            return "color: #22c55e; font-weight: bold" if float(val) >= 0 else "color: #ef4444; font-weight: bold"
                        except Exception:
                            return ""

                    pnl_styled = pnl_df.style.format({
                        "Sell Price ₹":    "{:.2f}",
                        "Sell Value ₹":    "{:,.2f}",
                        "Avg Buy Price ₹": lambda x: f"{x:.2f}" if pd.notna(x) else "—",
                        "Buy Cost ₹":      lambda x: f"{x:,.2f}" if pd.notna(x) else "—",
                        "Profit / Loss ₹": lambda x: f"₹{x:,.2f}" if pd.notna(x) else "—",
                        "Return %":        lambda x: f"{x:+.2f}%" if pd.notna(x) else "—",
                    }, na_rep="—")
                    if "Profit / Loss ₹" in pnl_df.columns:
                        pnl_styled = pnl_styled.map(_pnl_color, subset=["Profit / Loss ₹"])
                    if "Return %" in pnl_df.columns:
                        pnl_styled = pnl_styled.map(_pnl_color, subset=["Return %"])

                    st.dataframe(pnl_styled, use_container_width=True, hide_index=True)

                    # Summary metric
                    total_pnl = pnl_df["Profit / Loss ₹"].sum() if pnl_df["Profit / Loss ₹"].notna().any() else None
                    total_sold = pnl_df["Sell Value ₹"].sum() if "Sell Value ₹" in pnl_df.columns else None
                    mc1, mc2 = st.columns(2)
                    if total_sold:
                        mc1.metric("Total Sell Value ₹", f"{total_sold:,.2f}")
                    if total_pnl is not None:
                        mc2.metric("Total Realised P&L ₹", f"{total_pnl:,.2f}", delta=f"{total_pnl:+.2f}")

                    if not avg_buy_map:
                        st.info("No avg buy prices found in stock_config.xlsx. Go to **Trading Bot → Sync Holdings** to populate them, or enter them manually in the Excel sheet.")

    # ── AO Sub-tab 3: Order Book ───────────────────────────────────────────────
    with ao3:
        st.markdown("#### Today's Orders")
        st.caption("Includes pending, executed, cancelled, and rejected orders.")
        orders_res = st.session_state["ao_orders"]
        if not _ao_error(orders_res, "Order Book"):
            data = orders_res["data"]
            if not data:
                st.info("No orders placed today.")
            else:
                ORDER_COLS = {
                    "tradingsymbol":  "Symbol",
                    "exchange":       "Exchange",
                    "transactiontype":"Side",
                    "producttype":    "Product",
                    "quantity":       "Qty",
                    "price":          "Price ₹",
                    "orderstatus":    "Status",
                    "text":           "Reason",
                    "orderid":        "Order ID",
                    "updatetime":     "Updated",
                }
                df = pd.DataFrame(data)
                present = {k: v for k, v in ORDER_COLS.items() if k in df.columns}
                df = df[list(present.keys())].rename(columns=present)

                def _color_status(val):
                    v = str(val).upper()
                    if "COMPLETE" in v: return "color: #22c55e"
                    if "REJECT" in v or "CANCEL" in v: return "color: #ef4444"
                    if "OPEN" in v or "PENDING" in v:  return "color: #f59e0b"
                    return ""

                styled = df.style
                if "Status" in df.columns:
                    styled = styled.map(_color_status, subset=["Status"])

                st.dataframe(styled, use_container_width=True, hide_index=True)

                # Summary badge row
                statuses = [str(r.get("orderstatus", "")).upper() for r in data]
                n_done    = sum(1 for s in statuses if "COMPLETE" in s)
                n_open    = sum(1 for s in statuses if "OPEN" in s or "PENDING" in s)
                n_cancel  = sum(1 for s in statuses if "CANCEL" in s or "REJECT" in s)
                c1, c2, c3 = st.columns(3)
                c1.metric("✅ Executed", n_done)
                c2.metric("⏳ Open / Pending", n_open)
                c3.metric("❌ Cancelled / Rejected", n_cancel)

    # ── AO Sub-tab 4: Holdings ─────────────────────────────────────────────────
    with ao4:
        st.markdown("#### Demat Holdings")
        st.caption("Long-term holdings held in your demat account.")
        holdings_res = st.session_state["ao_holdings"]
        if not _ao_error(holdings_res, "Holdings"):
            data = holdings_res["data"]
            if not data:
                st.info("No holdings found.")
            else:
                HOLD_COLS = {
                    "tradingsymbol": "Symbol",
                    "exchange":      "Exchange",
                    "quantity":      "Qty",
                    "averageprice":  "Avg Price ₹",
                    "ltp":           "LTP ₹",
                    "close":         "Prev Close ₹",
                    "pnl":           "P&L ₹",
                    "producttype":   "Product",
                    "isin":          "ISIN",
                }
                df = pd.DataFrame(data)
                present = {k: v for k, v in HOLD_COLS.items() if k in df.columns}
                df = df[list(present.keys())].rename(columns=present)

                # Coerce numeric
                for col in ["Avg Price ₹", "LTP ₹", "P&L ₹", "Prev Close ₹"]:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce")

                styled = df.style.format(
                    {c: "{:.2f}" for c in ["Avg Price ₹", "LTP ₹", "P&L ₹", "Prev Close ₹"] if c in df.columns}
                )
                if "P&L ₹" in df.columns:
                    styled = styled.map(_color_pnl, subset=["P&L ₹"])

                st.dataframe(styled, use_container_width=True, hide_index=True)

                # Portfolio summary
                if "P&L ₹" in df.columns:
                    total_pnl = df["P&L ₹"].sum()
                    winners   = (df["P&L ₹"] > 0).sum()
                    losers    = (df["P&L ₹"] < 0).sum()
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Total Stocks", len(df))
                    c2.metric("Total P&L ₹", f"{total_pnl:,.2f}", delta=f"{total_pnl:+.2f}")
                    c3.metric("🟢 Winners", int(winners))
                    c4.metric("🔴 Losers", int(losers))

    # ── AO Sub-tab 5: Positions ────────────────────────────────────────────────
    with ao5:
        st.markdown("#### Open Positions")
        st.caption("Intraday (MIS) and carry-forward (CNC/NRML) open positions.")
        positions_res = st.session_state["ao_positions"]
        if not _ao_error(positions_res, "Positions"):
            data = positions_res["data"]
            if not data:
                st.info("No open positions.")
            else:
                POS_COLS = {
                    "tradingsymbol": "Symbol",
                    "exchange":      "Exchange",
                    "producttype":   "Product",
                    "netqty":        "Net Qty",
                    "netprice":      "Avg Price ₹",
                    "ltp":           "LTP ₹",
                    "close":         "Prev Close ₹",
                    "unrealised":    "Unrealised P&L ₹",
                    "realised":      "Realised P&L ₹",
                    "pnl":           "Total P&L ₹",
                }
                df = pd.DataFrame(data)
                present = {k: v for k, v in POS_COLS.items() if k in df.columns}
                df = df[list(present.keys())].rename(columns=present)

                pnl_cols = [c for c in ["Unrealised P&L ₹", "Realised P&L ₹", "Total P&L ₹"] if c in df.columns]
                price_cols = [c for c in ["Avg Price ₹", "LTP ₹", "Prev Close ₹"] if c in df.columns]

                for col in pnl_cols + price_cols:
                    df[col] = pd.to_numeric(df[col], errors="coerce")

                styled = df.style.format(
                    {c: "{:.2f}" for c in pnl_cols + price_cols}
                )
                for col in pnl_cols:
                    styled = styled.map(_color_pnl, subset=[col])

                st.dataframe(styled, use_container_width=True, hide_index=True)

                if "Total P&L ₹" in df.columns:
                    total = df["Total P&L ₹"].sum()
                    st.metric("Total Open P&L ₹", f"{total:,.2f}", delta=f"{total:+.2f}")


# ── Tab 6: Trading Bot — Phase 1 ───────────────────────────────────────────────
with tab6:
    st.subheader("🤖 Trading Bot — Phase 1: Signal Dashboard")
    st.caption("Dip-Buy / Rise-Sell signals for your 20-stock watchlist")

    if not is_configured():
        st.error("Angel One credentials not configured. Fill in `.env` first.")
        st.stop()

    # ── Setup check ────────────────────────────────────────────────────────────
    import os
    excel_exists = os.path.exists("stock_config.xlsx")
    if not excel_exists:
        st.warning("**stock_config.xlsx not found.**")
        st.markdown("""
Run this once in the Codespace terminal to generate it:
```bash
pip install openpyxl
python create_stock_config.py
```
Then open the **My Holdings** sheet and fill in your **Qty** and **Avg Buy Price** for any stocks you already own.
        """)
        st.stop()

    # ── Helper: get live api object ────────────────────────────────────────────
    def _get_live_api():
        from api.angelone import fetch_all as _ao_fetch_all, _load_token, _import_smart_connect
        ao_result = _ao_fetch_all()
        if not ao_result["login"]["status"]:
            st.error(f"Angel One auth failed: {ao_result['login']['error']}")
            return None
        cached      = _load_token()
        SmartConnect = _import_smart_connect()
        return SmartConnect(
            api_key=__import__("os").getenv("ANGELONE_API_KEY", ""),
            access_token=cached["access_token"]  if cached else None,
            refresh_token=cached.get("refresh_token") if cached else None,
            feed_token=cached.get("feed_token")   if cached else None,
        )

    # ── Action buttons ─────────────────────────────────────────────────────────
    col_sync, col_fetch, col_info = st.columns([1, 1, 3])
    with col_sync:
        do_sync = st.button("🔄 Sync Holdings", key="phase1_sync",
                            help="Auto-fill My Holdings sheet from your Angel One demat account")
    with col_fetch:
        run_signals = st.button("📡 Fetch Signals", type="primary", key="phase1_fetch")
    with col_info:
        if "p1_fetched_at" in st.session_state:
            st.caption(f"Last fetched: {st.session_state['p1_fetched_at']}")

    # ── Sync holdings ──────────────────────────────────────────────────────────
    if do_sync:
        from trading.phase1 import sync_holdings_from_api
        with st.spinner("Fetching your holdings from Angel One…"):
            # Prefer holdings already fetched in Angel One tab (uses fresh session)
            ao_holdings_cached = st.session_state.get("ao_holdings")
            raw_holdings = (ao_holdings_cached or {}).get("data") or []

            if raw_holdings:
                # Use already-fetched data — no extra API call needed
                from trading.phase1 import load_holdings
                from trading.stock_master import STOCK_MAP
                import openpyxl

                synced, skipped = [], []
                for h in raw_holdings:
                    raw_sym = str(h.get("tradingsymbol") or "").upper().replace("-EQ", "").strip()
                    try:
                        qty = float(h.get("quantity") or h.get("realisedquantity") or 0)
                        avg = float(h.get("averageprice") or 0)
                    except (ValueError, TypeError):
                        qty, avg = 0, 0
                    if qty > 0 and avg > 0:
                        if raw_sym in STOCK_MAP:
                            synced.append({"symbol": raw_sym, "name": STOCK_MAP[raw_sym]["name"],
                                           "qty": int(qty), "avg_buy_price": round(avg, 2)})
                        else:
                            skipped.append(raw_sym)

                # Write to Excel
                try:
                    wb  = openpyxl.load_workbook("stock_config.xlsx")
                    ws  = wb["My Holdings"]
                    header = {str(ws.cell(1, c).value).strip().lower(): c
                              for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
                    sym_col = next((v for k, v in header.items() if "symbol" in k), None)
                    qty_col = next((v for k, v in header.items() if "qty" in k), None)
                    avg_col = next((v for k, v in header.items() if "avg" in k or "buy price" in k), None)
                    synced_map = {s["symbol"]: s for s in synced}
                    for row in range(2, ws.max_row + 1):
                        sym = str(ws.cell(row, sym_col).value or "").strip().upper()
                        if sym in synced_map:
                            ws.cell(row, qty_col).value = synced_map[sym]["qty"]
                            ws.cell(row, avg_col).value = synced_map[sym]["avg_buy_price"]
                    wb.save("stock_config.xlsx")
                    st.success(f"✅ Synced {len(synced)} stock(s) into stock_config.xlsx → My Holdings")
                    if synced:
                        sync_df = pd.DataFrame(synced)[["name", "symbol", "qty", "avg_buy_price"]]
                        sync_df.columns = ["Stock", "Symbol", "Qty", "Avg Buy Price ₹"]
                        st.dataframe(sync_df, use_container_width=True, hide_index=True)
                    if skipped:
                        st.info(f"Skipped (not in 20-stock list): {', '.join(skipped)}")
                except Exception as e:
                    st.error(f"Excel write error: {e}")
            else:
                # Fallback: try live API call
                api_obj = _get_live_api()
                if api_obj:
                    sync_result = sync_holdings_from_api(api_obj)
                    if sync_result["error"]:
                        st.error(f"Sync error: {sync_result['error']}")
                        st.info("💡 **Fix:** Go to the **Angel One** tab first, click **Fetch from Angel One**, then come back here and click Sync Holdings.")
                    else:
                        synced = sync_result["synced"]
                        st.success(f"✅ Synced {len(synced)} stock(s)")
                        if synced:
                            sync_df = pd.DataFrame(synced)[["name", "symbol", "qty", "avg_buy_price"]]
                            sync_df.columns = ["Stock", "Symbol", "Qty", "Avg Buy Price ₹"]
                            st.dataframe(sync_df, use_container_width=True, hide_index=True)
                        if sync_result["skipped"]:
                            st.info(f"Skipped: {', '.join(sync_result['skipped'])}")

    if run_signals:
        from trading.phase1 import fetch_signals

        with st.spinner("Connecting to Angel One and fetching live prices…"):
            api_obj = _get_live_api()
            if not api_obj:
                st.stop()
            result = fetch_signals(api_obj)

        st.session_state["p1_result"]     = result
        st.session_state["p1_fetched_at"] = result["fetched_at"]

    if "p1_result" not in st.session_state:
        st.info("Click **Fetch Signals** to load live prices for all 20 stocks.")
        st.stop()

    result = st.session_state["p1_result"]
    df_all = result["signals"]
    buys   = result["buy_signals"]
    sells  = result["sell_signals"]
    errors = result["errors"]

    # ── Summary metrics ────────────────────────────────────────────────────────
    n_buy  = len(buys)
    n_sell = len(sells)
    n_hold = len(df_all) - n_buy - n_sell
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📋 Stocks Tracked", len(df_all))
    c2.metric("📈 BUY signals",  n_buy,  delta=f"+{n_buy}"  if n_buy  else None)
    c3.metric("📉 SELL signals", n_sell, delta=f"-{n_sell}" if n_sell else None)
    c4.metric("⏸ HOLD",         n_hold)

    if errors:
        with st.expander(f"⚠️ {len(errors)} stocks had fetch errors"):
            for sym, err in errors:
                st.markdown(f"- **{sym}**: {err}")

    st.divider()

    # ── BUY signals ────────────────────────────────────────────────────────────
    if buys:
        st.markdown("### 📈 BUY Signals Today")
        st.caption("Stocks that have dipped past your buy threshold — ranked by dip size")
        buy_df = pd.DataFrame([{
            "Stock":           b["Stock"],
            "Symbol":          b["Symbol"],
            "LTP ₹":           b["LTP ₹"],
            "Change %":        b["Change %"],
            "Buy Trigger %":   b["Buy Trigger %"],
            "Suggested Qty":   b.get("suggested_qty", "—"),
            "Capital ₹":       b.get("allocated_capital", "—"),
        } for b in buys])

        def _color_buy(val):
            try:
                return "color: #22c55e; font-weight: bold" if float(val) < 0 else ""
            except Exception:
                return ""

        styled_buy = buy_df.style.map(_color_buy, subset=["Change %"])
        st.dataframe(styled_buy, use_container_width=True, hide_index=True)
    else:
        st.info("No BUY signals today — no stock has dipped past its threshold.")

    st.divider()

    # ── SELL signals ───────────────────────────────────────────────────────────
    if sells:
        st.markdown("### 📉 SELL Signals Today")
        st.caption("Stocks that have reached your profit target (based on avg buy price)")
        sell_df = pd.DataFrame([{
            "Stock":          s["Stock"],
            "Symbol":         s["Symbol"],
            "LTP ₹":          s["LTP ₹"],
            "Avg Buy ₹":      s.get("Avg Buy ₹", "—"),
            "Sell Target ₹":  s.get("Sell Target ₹", "—"),
            "Qty Held":       s.get("Qty Held", "—"),
            "Change %":       s["Change %"],
        } for s in sells])
        st.dataframe(sell_df.style, use_container_width=True, hide_index=True)
    else:
        st.info("No SELL signals today.")

    st.divider()

    # ── Full signal table ──────────────────────────────────────────────────────
    st.markdown("### 📋 All 20 Stocks — Full Signal Table")

    def _color_signal(val):
        if val == "BUY":  return "color: #22c55e; font-weight: bold"
        if val == "SELL": return "color: #ef4444; font-weight: bold"
        if val == "HOLD": return "color: #f59e0b"
        return "color: #6b7280"

    def _color_change(val):
        try:
            v = float(val)
            return "color: #22c55e" if v > 0 else "color: #ef4444" if v < 0 else ""
        except Exception:
            return ""

    display_cols = ["Stock", "Category", "LTP ₹", "Prev Close ₹", "Change %",
                    "Buy Trigger %", "Avg Buy ₹", "Sell Target ₹", "Signal"]
    show_df = df_all[[c for c in display_cols if c in df_all.columns]].copy()

    styled_all = (
        show_df.style
        .map(_color_signal, subset=["Signal"])
        .map(_color_change, subset=["Change %"])
        .format({
            "LTP ₹":         "{:.2f}",
            "Prev Close ₹":  "{:.2f}",
            "Change %":      "{:+.2f}%",
            "Avg Buy ₹":     lambda x: f"{x:.2f}" if pd.notna(x) else "—",
            "Sell Target ₹": lambda x: f"{x:.2f}" if pd.notna(x) else "—",
        }, na_rep="—")
    )
    st.dataframe(styled_all, use_container_width=True, hide_index=True)

    st.divider()

    # ── Phase 2: Sell engine (holdings-aware) ──────────────────────────────────
    st.markdown("### 💰 Phase 2 — Sell Decisions (Your Holdings)")
    st.caption("Tier-1: sell half at target %. Tier-2: sell rest at 1.5× target %.")

    from trading.phase2 import compute_sell_signals, sell_summary as _sell_summary

    ltp_map = {
        row["Symbol"]: row["LTP ₹"]
        for _, row in df_all.iterrows()
        if pd.notna(row.get("LTP ₹"))
    }
    df_sell_p2 = compute_sell_signals(ltp_map)

    if df_sell_p2.empty:
        st.info("No holdings found. Click **Sync Holdings** to pull your positions from Angel One.")
    else:
        stats = _sell_summary(df_sell_p2)
        sc1, sc2, sc3 = st.columns(3)
        sc1.metric("Holdings tracked",    len(df_sell_p2))
        sc2.metric("Sell actions today",  stats["sell_count"])
        sc3.metric("Sell value ₹",        f"{stats['total_sell_value']:,.0f}" if stats['total_sell_value'] else "—")

        def _action_color(val):
            if "SELL FULL"  in str(val): return "color:#ef4444; font-weight:bold"
            if "SELL HALF"  in str(val): return "color:#f97316; font-weight:bold"
            if "SELL ALL"   in str(val): return "color:#ef4444; font-weight:bold"
            return "color:#6b7280"

        def _gain_color(val):
            try:
                return "color:#22c55e" if float(val) > 0 else "color:#ef4444"
            except Exception:
                return ""

        sell_styled = df_sell_p2.style \
            .map(_action_color, subset=["Action"]) \
            .map(_gain_color,   subset=["Gain %"])
        st.dataframe(sell_styled, use_container_width=True, hide_index=True)

    st.divider()

    # ── Phase 3: Trade plan ────────────────────────────────────────────────────
    st.markdown("### 📐 Phase 3 — Today's Buy Trade Plan")
    st.caption("Capital allocated across top signals — sector-capped, budget-aware.")

    from trading.phase3 import build_trade_plan, plan_as_dataframe

    budget_input = st.number_input(
        "Deployable budget today ₹ (keep ₹30K reserve separately)",
        min_value=5000, max_value=500000, value=120000, step=5000,
        key="p3_budget",
    )
    trade_plan = build_trade_plan(buys, total_budget=float(budget_input))
    plan_df    = plan_as_dataframe(trade_plan["plan"])

    p3c1, p3c2, p3c3 = st.columns(3)
    p3c1.metric("Stocks to buy",   len(trade_plan["plan"]))
    p3c2.metric("Capital deployed ₹", f"{trade_plan['deployed']:,.0f}")
    p3c3.metric("Undeployed ₹",    f"{trade_plan['remaining']:,.0f}")

    if not plan_df.empty:
        st.dataframe(plan_df, use_container_width=True, hide_index=True)
        with st.expander("Skipped signals (and why)"):
            if trade_plan["skipped"]:
                for s in trade_plan["skipped"]:
                    st.markdown(f"- **{s['Stock']}**: {s.get('skip_reason','—')}")
            else:
                st.write("None skipped.")
    else:
        st.info("No BUY signals to allocate today.")

    st.divider()

    # ── Phase 4: Paper trade log ───────────────────────────────────────────────
    st.markdown("### 📓 Phase 4 — Paper Trade Log")
    st.caption("Log of all signal decisions (no real orders). Run the scheduler to populate this.")

    from trading.phase4 import read_log, daily_summary, run_once as _p4_run_once

    col_p4a, col_p4b = st.columns([1, 3])
    with col_p4a:
        run_paper = st.button("▶ Run One Paper Cycle", key="p4_run",
                              help="Simulate one 15-min cycle right now and log it")
    with col_p4b:
        st.caption("Or run the scheduler in terminal: `python -m trading.phase4`")

    if run_paper:
        with st.spinner("Running paper trade cycle…"):
            api_obj = _get_live_api()
            if api_obj:
                p4_result = _p4_run_once(api_obj)
                st.success(f"Cycle {p4_result['run_id']} logged {p4_result['log_rows']} rows.")

    log_df = read_log(n_days=7)
    if log_df.empty:
        st.info("No paper trade log yet. Click **Run One Paper Cycle** to start.")
    else:
        summary = daily_summary(log_df)
        if summary:
            d1, d2, d3, d4 = st.columns(4)
            d1.metric("Runs today",             summary.get("runs_today", 0))
            d2.metric("BUY signals today",      summary.get("buy_signals", 0))
            d3.metric("SELL signals today",     summary.get("sell_signals", 0))
            d4.metric("Hypothetical deployed ₹", f"{summary.get('hypothetical_deployed', 0):,.0f}")

        st.markdown("**Last 7 days — all decisions:**")
        show_log = log_df[log_df["action"] != "HOLD"].copy()  # hide HOLDs by default

        def _log_color(val):
            if val == "BUY":       return "color:#22c55e; font-weight:bold"
            if "SELL" in str(val): return "color:#ef4444; font-weight:bold"
            return ""

        if not show_log.empty:
            st.dataframe(
                show_log.style.map(_log_color, subset=["action"]),
                use_container_width=True, hide_index=True,
            )
        with st.expander("Show HOLD rows too"):
            st.dataframe(log_df, use_container_width=True, hide_index=True)

    st.divider()

    # ── Daily Excel Trade Journal ──────────────────────────────────────────────
    st.markdown("### 📒 Daily Trade Journal (Excel)")
    st.caption("One sheet per trading day in `logs/paper_trades.xlsx` — auto-updated on every paper cycle.")

    from trading.excel_logger import (
        EXCEL_FILE as _EXCEL_FILE,
        get_all_sheet_names as _get_sheets,
        read_day_as_df as _read_day,
    )

    sheets = _get_sheets()
    if not sheets:
        st.info("No Excel journal yet. Click **▶ Run One Paper Cycle** above to create the first day's sheet.")
    else:
        # Download button
        try:
            with open(_EXCEL_FILE, "rb") as _f:
                st.download_button(
                    label="⬇ Download paper_trades.xlsx",
                    data=_f.read(),
                    file_name="paper_trades.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel",
                )
        except FileNotFoundError:
            pass

        # Day picker
        selected_day = st.selectbox("View day:", sheets, key="excel_day_picker")
        day_df = _read_day(selected_day)

        if day_df.empty:
            st.info("No data for this day.")
        else:
            # Split trade rows from summary rows (summary has no Time value)
            trade_mask = day_df["Time"].notna() & (day_df["Time"] != "")
            trade_view = day_df[trade_mask].copy()
            summ_view  = day_df[~trade_mask].dropna(how="all")

            def _ex_action_color(val):
                v = str(val).upper()
                if "BUY"  in v: return "color:#22c55e; font-weight:bold"
                if "SELL" in v: return "color:#ef4444; font-weight:bold"
                if "HOLD" in v: return "color:#f59e0b"
                return ""

            def _ex_pnl_color(val):
                try:
                    return "color:#22c55e; font-weight:bold" if float(str(val).replace(",","")) >= 0 else "color:#ef4444; font-weight:bold"
                except Exception:
                    return ""

            if not trade_view.empty:
                styled_ex = trade_view.style
                if "Action" in trade_view.columns:
                    styled_ex = styled_ex.map(_ex_action_color, subset=["Action"])
                if "Profit / Loss ₹" in trade_view.columns:
                    styled_ex = styled_ex.map(_ex_pnl_color, subset=["Profit / Loss ₹"])
                st.dataframe(styled_ex, use_container_width=True, hide_index=True)

            if not summ_view.empty:
                with st.expander("Day Summary"):
                    st.dataframe(summ_view[["Time", "Symbol"]].rename(
                        columns={"Time": "Metric", "Symbol": "Value"}
                    ), use_container_width=True, hide_index=True)

    st.divider()

    # ── Phase 5 preview: Live WebSocket Stream ─────────────────────────────────
    st.markdown("### 📡 Live Price Stream (WebSocket)")
    st.caption("Real-time LTP for all 20 stocks via Angel One SmartAPI Streaming 2.0")

    from trading import websocket_stream as _ws
    from trading.stock_master import STOCK_LIST, lookup_tokens

    col_ws1, col_ws2, col_ws3 = st.columns([1, 1, 3])

    with col_ws1:
        start_ws = st.button(
            "▶ Start Stream",
            key="ws_start",
            disabled=_ws.is_running(),
            help="Connect to Angel One WebSocket and start receiving live prices",
        )
    with col_ws2:
        stop_ws = st.button(
            "⏹ Stop Stream",
            key="ws_stop",
            disabled=not _ws.is_running(),
            help="Disconnect the WebSocket stream",
        )
    with col_ws3:
        stream_status = "🟢 Connected" if _ws.is_running() else "⚫ Disconnected"
        ticks = len(_ws.ltp_cache)
        st.caption(f"Status: **{stream_status}** &nbsp;·&nbsp; Symbols with live tick: **{ticks} / {len(STOCK_LIST)}**")

    if start_ws:
        with st.spinner("Connecting to Angel One WebSocket…"):
            from api.angelone import _load_token
            import os as _os

            token_data = _load_token()
            if not token_data:
                st.error("No cached Angel One token. Go to the **Angel One** tab and click **Fetch from Angel One** first.")
            else:
                jwt_token   = token_data.get("access_token", "")
                feed_token  = token_data.get("feed_token", "")
                api_key     = _os.getenv("ANGELONE_API_KEY", "")
                client_code = _os.getenv("ANGELONE_CLIENT_CODE", "")

                api_obj_ws = _get_live_api()
                if api_obj_ws:
                    with st.spinner("Looking up token IDs for 20 stocks…"):
                        token_map = lookup_tokens(api_obj_ws)

                    missing = [s["symbol"] for s in STOCK_LIST if not token_map.get(s["symbol"], {}).get("token")]
                    if missing:
                        st.warning(f"Token IDs missing for: {', '.join(missing)}. They will be skipped.")

                    ok = _ws.start_stream(jwt_token, api_key, client_code, feed_token, token_map)
                    if ok:
                        st.success("Stream started — prices will update in the table below. Click **Refresh** to see latest values.")
                    else:
                        st.error("Could not start stream. Check logs for details.")

    if stop_ws:
        _ws.stop_stream()
        st.info("Stream disconnected.")

    # ── Live price table ───────────────────────────────────────────────────────
    col_ref, _ = st.columns([1, 4])
    with col_ref:
        refresh_ws = st.button("🔄 Refresh prices", key="ws_refresh")

    # Build display table: all 20 stocks, fill in streamed LTP where available
    all_live = _ws.get_all_ltp()
    ws_rows = []
    for stock in STOCK_LIST:
        sym  = stock["symbol"]
        live = all_live.get(sym)
        entry = _ws.ltp_cache.get(sym)
        try:
            from datetime import datetime as _dt
            ts = entry["ts"] if entry else None
            updated = _dt.fromtimestamp(ts / 1000).strftime("%H:%M:%S") if ts else "—"
        except Exception:
            updated = "—"
        ws_rows.append({
            "Stock":        stock["name"],
            "Symbol":       sym,
            "Category":     stock["category"],
            "Live LTP ₹":  round(live, 2) if live is not None else None,
            "Last Tick":    updated,
        })

    ws_df = pd.DataFrame(ws_rows)

    def _ws_ltp_color(val):
        if val is None or str(val) == "None":
            return "color: #6b7280"
        return "color: #22c55e; font-weight: bold"

    ws_styled = ws_df.style.map(_ws_ltp_color, subset=["Live LTP ₹"])
    st.dataframe(ws_styled, use_container_width=True, hide_index=True)

    if not _ws.is_running() and not all_live:
        st.info("Click **▶ Start Stream** to begin receiving live prices.")

    st.divider()

    # ── Phase 5: Paper Portfolio Tracker ───────────────────────────────────────
    st.markdown("### 🗂 Phase 5 — Paper Portfolio Tracker")
    st.caption("Simulates a virtual ₹1.5L portfolio built from your paper trade log — tracks open positions and realised P&L.")

    from trading.phase5 import build_portfolio as _build_portfolio

    p5_budget = st.number_input(
        "Starting virtual cash ₹",
        min_value=10000, max_value=1000000, value=150000, step=10000,
        key="p5_budget",
    )

    # Use streamed LTPs if available, else fall back to Phase 1 signals
    p5_ltp = dict(_ws.get_all_ltp()) if _ws.get_all_ltp() else {}
    if not p5_ltp and "p1_result" in st.session_state:
        for _, row in st.session_state["p1_result"]["signals"].iterrows():
            if pd.notna(row.get("LTP ₹")):
                p5_ltp[row["Symbol"]] = row["LTP ₹"]

    portfolio = _build_portfolio(starting_cash=float(p5_budget), ltp_map=p5_ltp or None)

    if portfolio["log_df"].empty:
        st.info("No paper trade log yet. Run Phase 4 cycles to populate it.")
    else:
        # Summary metrics
        total_gain = portfolio["realised_pnl"] + portfolio["unrealised_pnl"]
        pm1, pm2, pm3, pm4, pm5 = st.columns(5)
        pm1.metric("Portfolio Value ₹",   f"{portfolio['portfolio_value']:,.0f}")
        pm2.metric("Available Cash ₹",    f"{portfolio['cash']:,.0f}")
        pm3.metric("Realised P&L ₹",      f"{portfolio['realised_pnl']:+,.0f}",
                   delta=f"{portfolio['realised_pnl']:+.0f}")
        pm4.metric("Unrealised P&L ₹",    f"{portfolio['unrealised_pnl']:+,.0f}"
                                           if p5_ltp else "— (no LTP)",
                   delta=f"{portfolio['unrealised_pnl']:+.0f}" if p5_ltp else None)
        pm5.metric("Win / Loss",
                   f"{portfolio['win_count']}W / {portfolio['loss_count']}L")

        st.markdown("#### Open Paper Positions")
        open_df = portfolio["open_positions"]
        if open_df.empty:
            st.info("No open paper positions.")
        else:
            def _unreal_color(val):
                try:
                    return "color:#22c55e" if float(val) >= 0 else "color:#ef4444"
                except Exception:
                    return "color:#6b7280"

            open_styled = open_df.style.format({
                "Entry Price ₹":    "{:.2f}",
                "LTP ₹":            lambda x: f"{x:.2f}" if pd.notna(x) else "—",
                "Current Value ₹":  lambda x: f"₹{x:,.2f}" if pd.notna(x) else "—",
                "Unrealised P&L ₹": lambda x: f"₹{x:+,.2f}" if pd.notna(x) else "—",
                "Return %":         lambda x: f"{x:+.2f}%" if pd.notna(x) else "—",
            }, na_rep="—")
            if "Unrealised P&L ₹" in open_df.columns:
                open_styled = open_styled.map(_unreal_color, subset=["Unrealised P&L ₹"])
            st.dataframe(open_styled, use_container_width=True, hide_index=True)

        st.markdown("#### Closed Paper Trades")
        closed_df = portfolio["closed_trades"]
        if closed_df.empty:
            st.info("No closed paper trades yet.")
        else:
            closed_styled = closed_df.style.format({
                "Entry Price ₹":    "{:.2f}",
                "Exit Price ₹":     "{:.2f}",
                "Realised P&L ₹":  "₹{:+,.2f}",
                "Return %":         "{:+.2f}%",
            })
            closed_styled = closed_styled.map(_pnl_color, subset=["Realised P&L ₹"])
            st.dataframe(closed_styled, use_container_width=True, hide_index=True)

    st.divider()

    # ── Phase 6: Strategy Performance Analytics ────────────────────────────────
    st.markdown("### 📊 Phase 6 — Strategy Performance Analytics")
    st.caption("Win rate, category edge, P&L curve — validates the algorithm before going live.")

    from trading.phase6 import (
        signal_summary as _sig_summary,
        closed_trade_stats as _trade_stats,
        category_breakdown as _cat_breakdown,
        daily_pnl_curve as _pnl_curve,
        top_signals as _top_signals,
        signal_accuracy as _sig_accuracy,
    )

    if portfolio["log_df"].empty:
        st.info("No data yet — run Phase 4 paper trading cycles first.")
    else:
        log_df_p6 = portfolio["log_df"]
        closed_p6 = portfolio["closed_trades"]

        # ── Signal summary ─────────────────────────────────────────────────────
        sig = _sig_summary(log_df_p6)
        if sig:
            sa1, sa2, sa3, sa4, sa5 = st.columns(5)
            sa1.metric("Total Runs",      sig.get("total_runs", 0))
            sa2.metric("BUY Signals",     sig.get("buy_signals", 0))
            sa3.metric("SELL Signals",    sig.get("sell_signals", 0))
            sa4.metric("Total Deployed ₹", f"{sig.get('total_deployed', 0):,.0f}")
            sa5.metric("Date Range",      sig.get("date_range", "—"))

        st.markdown("#### Trade Performance")
        if not closed_p6.empty:
            stats = _trade_stats(closed_p6)
            tc1, tc2, tc3, tc4, tc5 = st.columns(5)
            tc1.metric("Closed Trades",  stats.get("total_trades", 0))
            tc2.metric("Win Rate",       f"{stats.get('win_rate_pct', 0):.1f}%")
            tc3.metric("Avg Return",     f"{stats.get('avg_return_pct', 0):+.2f}%")
            tc4.metric("Total Realised ₹", f"{stats.get('total_realised', 0):+,.0f}")
            tc5.metric("Profit Factor",  stats.get("profit_factor", "—"))

            # Best / worst trades
            bc1, bc2 = st.columns(2)
            best  = stats.get("best_trade", {})
            worst = stats.get("worst_trade", {})
            with bc1:
                if best:
                    st.success(f"**Best trade:** {best.get('Stock','—')} — "
                               f"₹{best.get('Realised P&L ₹',0):+,.0f} "
                               f"({best.get('Return %',0):+.2f}%)")
            with bc2:
                if worst:
                    st.error(f"**Worst trade:** {worst.get('Stock','—')} — "
                             f"₹{worst.get('Realised P&L ₹',0):+,.0f} "
                             f"({worst.get('Return %',0):+.2f}%)")

            # Top / bottom signals
            top5, bot5 = _top_signals(closed_p6)
            t1, t2 = st.columns(2)
            with t1:
                st.markdown("**Top 5 trades**")
                st.dataframe(top5[["Stock", "Return %", "Realised P&L ₹"]].style.format(
                    {"Return %": "{:+.2f}%", "Realised P&L ₹": "₹{:+,.2f}"}
                ), use_container_width=True, hide_index=True)
            with t2:
                st.markdown("**Bottom 5 trades**")
                st.dataframe(bot5[["Stock", "Return %", "Realised P&L ₹"]].style.format(
                    {"Return %": "{:+.2f}%", "Realised P&L ₹": "₹{:+,.2f}"}
                ), use_container_width=True, hide_index=True)
        else:
            st.info("No closed paper trades yet to analyse.")

        # ── Category breakdown ─────────────────────────────────────────────────
        st.markdown("#### Category / Sector Edge")
        cat_df = _cat_breakdown(log_df_p6, closed_p6 if not closed_p6.empty else None)
        if not cat_df.empty:
            st.dataframe(cat_df, use_container_width=True, hide_index=True)

        # ── P&L curve ─────────────────────────────────────────────────────────
        if not closed_p6.empty:
            pnl_curve = _pnl_curve(closed_p6)
            if not pnl_curve.empty:
                st.markdown("#### Cumulative P&L Curve")
                st.line_chart(
                    pnl_curve.set_index("Date")[["Cumulative P&L ₹"]],
                    use_container_width=True,
                )

        # ── Signal accuracy vs current LTP ────────────────────────────────────
        if p5_ltp:
            st.markdown("#### Signal Accuracy (vs current LTP)")
            acc_df = _sig_accuracy(log_df_p6, p5_ltp)
            if not acc_df.empty:
                def _acc_color(val):
                    return "color:#22c55e" if "Yes" in str(val) else "color:#ef4444"
                st.dataframe(
                    acc_df.style.map(_acc_color, subset=["Correct?"]).format({
                        "Signal Price ₹":       "{:.2f}",
                        "Current LTP ₹":        "{:.2f}",
                        "Gain Since Signal %":  "{:+.2f}%",
                    }),
                    use_container_width=True, hide_index=True,
                )
