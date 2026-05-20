"""
Excel Daily Trade Logger

Writes paper trading activity to logs/paper_trades.xlsx.
Each trading day gets its own sheet named "DD-Mon-YYYY".

Sheet layout:
  Row 1         : Header (bold, green)
  Rows 2-N      : One row per signal (BUY / SELL / HOLD)
  Row N+2       : Blank separator
  Row N+3 onward: Day summary block (total deployed, P&L, etc.)

Columns:
  Time | Symbol | Stock | Category | Action | Signal Price ₹ |
  Qty  | Amount ₹ | Avg Buy Price ₹ | Buy Cost ₹ |
  Profit/Loss ₹ | Return % | Reason
"""

import os
from datetime import datetime, date
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

EXCEL_DIR  = "logs"
EXCEL_FILE = os.path.join(EXCEL_DIR, "paper_trades.xlsx")

# Colours
_GREEN_FILL  = PatternFill("solid", fgColor="1F6B35")
_BUY_FILL    = PatternFill("solid", fgColor="16A34A")   # green
_SELL_FILL   = PatternFill("solid", fgColor="DC2626")   # red
_HOLD_FILL   = PatternFill("solid", fgColor="92400E")   # amber
_SUMM_FILL   = PatternFill("solid", fgColor="1E3A5F")   # dark blue
_PROFIT_FONT = Font(color="22C55E", bold=True)
_LOSS_FONT   = Font(color="EF4444", bold=True)
_WHITE_BOLD  = Font(color="FFFFFF", bold=True)
_WHITE       = Font(color="FFFFFF")

HEADERS = [
    "Time", "Symbol", "Stock", "Category",
    "Action", "Signal Price ₹", "Qty", "Amount ₹",
    "Avg Buy Price ₹", "Buy Cost ₹", "Profit / Loss ₹", "Return %",
    "Reason",
]

COL_WIDTHS = [10, 14, 22, 14, 14, 16, 6, 14, 16, 14, 16, 10, 40]


def _sheet_name(for_date: date | str) -> str:
    if isinstance(for_date, str):
        for_date = datetime.strptime(for_date[:10], "%Y-%m-%d").date()
    return for_date.strftime("%d-%b-%Y")


def _ensure_workbook() -> openpyxl.Workbook:
    os.makedirs(EXCEL_DIR, exist_ok=True)
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet
    return wb


def _write_header(ws, row: int = 1):
    for col, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.fill      = _GREEN_FILL
        cell.font      = _WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 22


def _action_fill(action: str) -> PatternFill | None:
    a = action.upper()
    if "BUY"  in a: return _BUY_FILL
    if "SELL" in a: return _SELL_FILL
    if "HOLD" in a: return _HOLD_FILL
    return None


def _thin_border() -> Border:
    s = Side(style="thin", color="444444")
    return Border(left=s, right=s, top=s, bottom=s)


def write_day(
    trade_rows: list[dict],
    avg_buy_map: dict[str, float] | None = None,
    for_date: date | str | None = None,
):
    """
    Write (or overwrite) the sheet for a given trading day.

    Args:
        trade_rows  : list of dicts from phase4.run_once() log_rows
                      keys: date, time, symbol, stock, category,
                            action, signal_price, qty, amount, reason
        avg_buy_map : {symbol: avg_buy_price} — used to compute P&L for SELL rows
        for_date    : date to use; defaults to today
    """
    if for_date is None:
        for_date = date.today()
    sheet_name = _sheet_name(for_date)

    wb = _ensure_workbook()
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    _write_header(ws, row=1)

    # ── Trade rows ─────────────────────────────────────────────────────────────
    total_deployed = 0.0
    total_sell_val = 0.0
    total_pnl      = 0.0
    buy_count = sell_count = hold_count = 0

    border = _thin_border()

    for r_idx, row in enumerate(trade_rows, start=2):
        action     = str(row.get("action", "")).upper()
        sym        = str(row.get("symbol", "")).upper()
        qty        = row.get("qty", 0)
        price      = row.get("signal_price", "")
        amount     = row.get("amount", 0)
        reason     = str(row.get("reason", ""))

        # P&L for SELL rows
        avg_buy  = (avg_buy_map or {}).get(sym)
        buy_cost = pnl = ret_pct = None
        if "SELL" in action and avg_buy and qty and price:
            try:
                buy_cost = round(float(qty) * avg_buy, 2)
                pnl      = round(float(amount) - buy_cost, 2)
                ret_pct  = round(pnl / buy_cost * 100, 2) if buy_cost else None
                total_pnl += pnl
            except (TypeError, ValueError):
                pass

        if "BUY"  in action: buy_count  += 1; total_deployed += float(amount or 0)
        if "SELL" in action: sell_count += 1; total_sell_val += float(amount or 0)
        if "HOLD" in action: hold_count += 1

        row_vals = [
            row.get("time", ""),
            sym,
            row.get("stock", ""),
            row.get("category", ""),
            action,
            price if price else "",
            int(qty) if qty else 0,
            round(float(amount), 2) if amount else 0,
            round(avg_buy, 2)  if avg_buy  else "",
            buy_cost           if buy_cost  is not None else "",
            pnl                if pnl       is not None else "",
            f"{ret_pct:+.2f}%" if ret_pct  is not None else "",
            reason,
        ]

        fill = _action_fill(action)
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
                cell.font = _WHITE

        # Colour P&L column
        pnl_cell = ws.cell(row=r_idx, column=11)
        if pnl is not None:
            pnl_cell.font = _PROFIT_FONT if pnl >= 0 else _LOSS_FONT

    # ── Day summary block ──────────────────────────────────────────────────────
    sep_row  = len(trade_rows) + 3
    summ_row = sep_row

    summary = [
        ("BUY signals",          buy_count),
        ("SELL signals",         sell_count),
        ("HOLD",                 hold_count),
        ("Total deployed ₹",     round(total_deployed, 2)),
        ("Total sell value ₹",   round(total_sell_val, 2)),
        ("Realised P&L ₹",       round(total_pnl, 2)),
    ]
    for label, val in summary:
        lc = ws.cell(row=summ_row, column=1, value=label)
        vc = ws.cell(row=summ_row, column=2, value=val)
        lc.fill = vc.fill = _SUMM_FILL
        lc.font = vc.font = _WHITE_BOLD
        lc.alignment = Alignment(horizontal="right")
        vc.alignment = Alignment(horizontal="left")
        # Colour P&L value
        if label == "Realised P&L ₹":
            vc.font = Font(
                color="22C55E" if total_pnl >= 0 else "EF4444",
                bold=True,
            )
        summ_row += 1

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(EXCEL_FILE)
    return EXCEL_FILE


def get_all_sheet_names() -> list[str]:
    """Return list of day-sheet names in the workbook (newest first)."""
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    names = list(reversed(wb.sheetnames))
    wb.close()
    return names


def read_day_as_df(sheet_name: str) -> "pd.DataFrame":
    """Read one day's sheet back as a DataFrame."""
    import pandas as pd
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = list(ws.values)
    wb.close()
    if len(rows) < 2:
        return pd.DataFrame()
    # Header is row 0; find data rows (stop at blank row = summary separator)
    headers = rows[0]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            break
        data.append(row)
    return pd.DataFrame(data, columns=headers)
