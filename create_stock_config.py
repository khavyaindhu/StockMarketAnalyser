"""
Run once to generate stock_config.xlsx.
  python create_stock_config.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STOCKS = [
    # ── Pharma ────────────────────────────────────────────────────────────────
    ("Cipla",               "CIPLA",       "NSE", "Pharma",            2.5, 5.0, 15000),
    ("Natco Pharma",        "NATCOPHARM",  "NSE", "Pharma",            2.5, 5.0, 15000),
    ("Dr Reddys",           "DRREDDY",     "NSE", "Pharma",            2.5, 5.0, 15000),
    ("Zydus Lifesciences",  "ZYDUSLIFE",   "NSE", "Pharma",            2.5, 5.0, 15000),
    # ── Large Cap / Quality ───────────────────────────────────────────────────
    ("HDFC Bank",           "HDFCBANK",    "NSE", "Large Cap",         2.0, 4.0, 15000),
    ("ITC",                 "ITC",         "NSE", "Large Cap",         2.0, 4.0, 15000),
    ("Wipro",               "WIPRO",       "NSE", "Large Cap",         2.0, 4.0, 15000),
    # ── Mid Banks ─────────────────────────────────────────────────────────────
    ("Federal Bank",        "FEDERALBNK",  "NSE", "Mid Bank",          3.0, 6.0, 12000),
    ("IDFC First Bank",     "IDFCFIRSTB",  "NSE", "Mid Bank",          3.0, 6.0, 12000),
    ("IndusInd Bank",       "INDUSINDBK",  "NSE", "Mid Bank",          3.0, 6.0, 12000),
    ("Karnataka Bank",      "KTKBANK",     "NSE", "Mid Bank",          3.0, 6.0, 12000),
    ("South Indian Bank",   "SOUTHBANK",   "NSE", "Mid Bank",          3.0, 6.0, 12000),
    # ── NBFC / Finance ────────────────────────────────────────────────────────
    ("Manappuram Finance",  "MANAPPURAM",  "NSE", "NBFC",              3.5, 7.0, 12000),
    # ── Metals / Cyclical ─────────────────────────────────────────────────────
    ("Tata Steel",          "TATASTEEL",   "NSE", "Metals",            3.5, 6.0, 12000),
    # ── Auto ──────────────────────────────────────────────────────────────────
    ("Tata Motors CV",      "TMCV",        "NSE", "Auto",              3.0, 6.0, 12000),
    ("Tata Motors PV",      "TMPV",        "NSE", "Auto",              3.0, 6.0, 12000),
    # ── Jewellery / Consumer ──────────────────────────────────────────────────
    ("ITC Hotels",          "ITCHOTELS",   "NSE", "Consumer",          3.0, 6.0, 12000),
    ("Kalyan Jewellers",    "KALYANKJIL",  "NSE", "Jewellery",         3.0, 6.0, 12000),
    ("Thangamayil",         "THANGAMAYL",  "NSE", "Jewellery",         3.0, 6.0, 12000),
    ("Titan",               "TITAN",       "NSE", "Jewellery",         3.0, 6.0, 12000),
]

HOLDINGS_TEMPLATE = [
    # stock_name, symbol, qty, avg_buy_price  ← user fills qty + avg_buy_price
    ("Cipla",               "CIPLA",       0, 0.0),
    ("Natco Pharma",        "NATCOPHARM",  0, 0.0),
    ("Dr Reddys",           "DRREDDY",     0, 0.0),
    ("Zydus Lifesciences",  "ZYDUSLIFE",   0, 0.0),
    ("HDFC Bank",           "HDFCBANK",    0, 0.0),
    ("ITC",                 "ITC",         0, 0.0),
    ("Wipro",               "WIPRO",       0, 0.0),
    ("Federal Bank",        "FEDERALBNK",  0, 0.0),
    ("IDFC First Bank",     "IDFCFIRSTB",  0, 0.0),
    ("IndusInd Bank",       "INDUSINDBK",  0, 0.0),
    ("Karnataka Bank",      "KTKBANK",     0, 0.0),
    ("South Indian Bank",   "SOUTHBANK",   0, 0.0),
    ("Manappuram Finance",  "MANAPPURAM",  0, 0.0),
    ("Tata Steel",          "TATASTEEL",   0, 0.0),
    ("Tata Motors CV",      "TMCV",        0, 0.0),
    ("Tata Motors PV",      "TMPV",        0, 0.0),
    ("ITC Hotels",          "ITCHOTELS",   0, 0.0),
    ("Kalyan Jewellers",    "KALYANKJIL",  0, 0.0),
    ("Thangamayil",         "THANGAMAYL",  0, 0.0),
    ("Titan",               "TITAN",       0, 0.0),
]

# ── Style helpers ──────────────────────────────────────────────────────────────

def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font():
    return Font(bold=True, color="FFFFFF", size=11)

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _write_header(ws, row, cols, fill_hex):
    fill   = _header_fill(fill_hex)
    font   = _header_font()
    border = _thin_border()
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill      = fill
        cell.font      = font
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _write_data_row(ws, row, values, alt=False):
    bg  = "F0F4FF" if alt else "FFFFFF"
    fill   = PatternFill("solid", fgColor=bg)
    border = _thin_border()
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Stock Config ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Stock Config"
    ws1.row_dimensions[1].height = 28

    headers = [
        "Stock Name", "NSE Symbol", "Exchange", "Category",
        "Buy Dip %\n(from prev close)", "Sell Target %\n(from avg buy price)",
        "Max Capital ₹\n(per trade)", "Token ID\n(auto-filled)", "Active",
    ]
    _write_header(ws1, 1, headers, "1E3A5F")

    widths = [22, 16, 12, 14, 18, 22, 18, 18, 10]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws1, i, w)

    for r, (name, sym, exch, cat, buy, sell, cap) in enumerate(STOCKS, 2):
        _write_data_row(ws1, r, [name, sym, exch, cat, buy, sell, cap, "", "YES"], alt=(r % 2 == 0))

    # ── Sheet 2: My Holdings ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("My Holdings")
    ws2.row_dimensions[1].height = 28

    h2 = [
        "Stock Name", "NSE Symbol",
        "Qty Held\n(fill in)", "Avg Buy Price ₹\n(fill in)",
        "Invested ₹\n(auto)", "Sell Target Price ₹\n(auto)",
        "Notes",
    ]
    _write_header(ws2, 1, h2, "1B5E20")

    w2 = [22, 16, 14, 20, 16, 22, 24]
    for i, w in enumerate(w2, 1):
        _set_col_width(ws2, i, w)

    # find sell_target_pct by symbol lookup
    sell_map = {sym: sell for _, sym, _, _, _, sell, _ in STOCKS}

    for r, (name, sym, qty, avg) in enumerate(HOLDINGS_TEMPLATE, 2):
        sp = sell_map.get(sym, 5.0)
        # Invested = qty * avg  (formula)
        invested_formula    = f"=C{r}*D{r}"
        # Sell target price = avg * (1 + sell_target_pct/100)
        sell_price_formula  = f"=IF(D{r}>0, D{r}*(1+{sp}/100), \"\")"
        _write_data_row(ws2, r, [name, sym, qty or "", avg or "",
                                  invested_formula, sell_price_formula, ""],
                        alt=(r % 2 == 0))
        # Qty and Avg Buy Price cells — highlight yellow for user to fill
        for col in (3, 4):
            cell = ws2.cell(row=r, column=col)
            cell.fill = PatternFill("solid", fgColor="FFF9C4")
            cell.font = Font(bold=True, size=11)

    # ── Sheet 3: Trade Log ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Trade Log")
    ws3.row_dimensions[1].height = 28
    h3 = [
        "Date", "Time", "Stock", "Symbol", "Action",
        "Qty", "Price ₹", "Amount ₹", "Reason", "Order ID", "Status",
    ]
    _write_header(ws3, 1, h3, "4A148C")
    w3 = [14, 10, 20, 14, 10, 8, 12, 14, 28, 18, 12]
    for i, w in enumerate(w3, 1):
        _set_col_width(ws3, i, w)

    wb.save("stock_config.xlsx")
    print("✅  stock_config.xlsx created successfully.")
    print("   → Open 'My Holdings' sheet and fill in Qty + Avg Buy Price for stocks you already own.")


if __name__ == "__main__":
    build()
